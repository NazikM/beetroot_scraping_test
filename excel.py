from abc import ABC, abstractmethod
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define font and fill styles
bold_font = Font(bold=True)
blue_fill = PatternFill(start_color="A7C4E5", end_color="A7C4E5", fill_type="solid")
# Define border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


class FileWriter(ABC):
    def __init__(self, file_name):
        self.file_name = file_name

    @abstractmethod
    def write_data(self, data):
        pass

    @abstractmethod
    def save(self):
        pass


class ExcelWriter(FileWriter):
    def __init__(self, file_name):
        super().__init__(file_name)
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def write_header(self):
        def set_cell_properties(value):
            cell.font = bold_font
            cell.fill = blue_fill
            cell.value = value
            cell.border = thin_border

        # Set column widths for better alignment
        column_widths = [30, 30, 30, 30, 30, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            self.sheet.column_dimensions[chr(64 + col_idx)].width = width

        content = ["Name (incl. titles if any mentioned)", "Affiliation(s) Name(s)", "Person's Location", "Session Name",
             "Topic Title", "Presentation Abstract"]

        cell = self.sheet.cell(row=1, column=1)
        set_cell_properties("About the person")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = self.sheet.cell(row=1, column=4)
        set_cell_properties("About the session/topic")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        for col, val in enumerate(content, start=1):
            cell = self.sheet.cell(row=2, column=col)
            set_cell_properties(val)
        # Merge cells in the header rows
        self.sheet.merge_cells("A1:C1")
        self.sheet.merge_cells("D1:F1")

    def write_data(self, data):
        self.write_header()
        for row, entry in enumerate(data, start=3):
            for col, val in enumerate(entry, start=1):
                self.sheet.cell(row=row, column=col).value = val

    def save(self):
        self.workbook.save(self.file_name)


if __name__ == "__main__":
    data_to_write = [
        ["Melinda Gooderham, Boni E Elewski", "Probity Medical Research, Waterloo", "Kenilworth, USA", "P052", "INCIDENCE OF SERIOUS GASTROINTESTINAL EVENTS AMONG", "In pts treated with IXE, improvements in the signs and\tsymptoms\tof"],
    ]

    filename = "temp.xlsx"
    excel_writer = ExcelWriter(filename)
    excel_writer.write_data(data_to_write)
    excel_writer.save()
